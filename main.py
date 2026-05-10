"""
Tesouro Direto Real Value Calculator

Parses all Tesouro Direto analytical extract XLSX files from the reports/
directory, fetches the current IPCA rate from Brasil API, and calculates
the real accrued value of each investment based on the contracted rate
(IPCA + spread), comparing it with the market value shown in the report.
"""

import io
import os
import re
import sys
from datetime import date, datetime, timedelta

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
BRASIL_API_URL = "https://brasilapi.com.br/api/taxas/v1"
HOLIDAYS_API_URL = "https://brasilapi.com.br/api/feriados/v1"


def fetch_ipca_rate():
    """Fetch the current yearly IPCA rate from Brasil API."""
    resp = requests.get(BRASIL_API_URL, timeout=10)
    resp.raise_for_status()
    for entry in resp.json():
        if entry["nome"].upper() == "IPCA":
            return entry["valor"] / 100
    raise ValueError("IPCA rate not found in Brasil API response")


def fetch_holidays(year_start, year_end):
    """Fetch Brazilian national holidays for a range of years.

    Returns a set of date objects.
    """
    holidays = set()
    for year in range(year_start, year_end + 1):
        resp = requests.get(f"{HOLIDAYS_API_URL}/{year}", timeout=10)
        resp.raise_for_status()
        for entry in resp.json():
            holidays.add(datetime.strptime(entry["date"], "%Y-%m-%d").date())
    return holidays


def count_business_days(start, end, holidays):
    """Count business days (dias úteis) between start (exclusive) and end (inclusive).

    Excludes weekends (Sat=5, Sun=6) and national holidays.
    """
    count = 0
    d = start + timedelta(days=1)
    while d <= end:
        if d.weekday() < 5 and d not in holidays:
            count += 1
        d += timedelta(days=1)
    return count


def parse_br_number(value):
    """Parse a Brazilian-formatted number (e.g. '3.197,24' or '3,27')."""
    s = str(value).strip()
    # If it's already a plain float from openpyxl, just convert
    if isinstance(value, (int, float)):
        return float(value)
    # Brazilian format: dots are thousands separators, comma is decimal
    s = s.replace(".", "").replace(",", ".")
    return float(s)


def parse_xlsx(filepath):
    """Parse a single Tesouro Direto analytical extract XLSX file.

    Returns a dict with bond metadata and a list of purchase entries.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Row 1: bond name
    bond_name = ws.cell(1, 1).value
    bond_name = bond_name.replace("EXTRATO ANALÍTICO - ", "").strip()

    # Row 4: maturity
    maturity_str = ws.cell(4, 1).value  # "VENCIMENTO: 15/12/2059"
    maturity_date_str = maturity_str.split(":")[1].strip()
    maturity_date = datetime.strptime(maturity_date_str, "%d/%m/%Y").date()

    # Data rows start at row 8, until we hit "Total"
    purchases = []
    row_idx = 8
    while True:
        cell_a = ws.cell(row_idx, 1).value
        if cell_a is None or str(cell_a).strip().lower() == "total":
            break

        purchase_date = (
            datetime.strptime(cell_a, "%d/%m/%Y").date()
            if isinstance(cell_a, str)
            else cell_a.date()
            if isinstance(cell_a, datetime)
            else cell_a
        )

        qty = parse_br_number(ws.cell(row_idx, 2).value)
        price = parse_br_number(ws.cell(row_idx, 3).value)
        invested = parse_br_number(ws.cell(row_idx, 4).value)

        # Contracted rate: "IPCA + 6,38%"
        rate_str = str(ws.cell(row_idx, 5).value)
        spread_match = re.search(r"[\d,.]+", rate_str.split("+")[-1])
        spread = parse_br_number(spread_match.group()) / 100

        # Market values from the report
        gross_value = parse_br_number(ws.cell(row_idx, 8).value)
        days = int(parse_br_number(ws.cell(row_idx, 9).value))
        ir_rate = parse_br_number(ws.cell(row_idx, 10).value) / 100
        ir_tax = parse_br_number(ws.cell(row_idx, 11).value)
        b3_fee = parse_br_number(ws.cell(row_idx, 13).value)
        net_value = parse_br_number(ws.cell(row_idx, 15).value)

        purchases.append(
            {
                "date": purchase_date,
                "qty": qty,
                "price": price,
                "invested": invested,
                "spread": spread,
                "mkt_gross": gross_value,
                "days": days,
                "ir_rate": ir_rate,
                "ir_tax": ir_tax,
                "b3_fee": b3_fee,
                "mkt_net": net_value,
            }
        )
        row_idx += 1

    # Total row
    total_row = row_idx
    total_invested = parse_br_number(ws.cell(total_row, 4).value)
    total_mkt_gross = parse_br_number(ws.cell(total_row, 8).value)
    total_mkt_net = parse_br_number(ws.cell(total_row, 15).value)

    wb.close()

    return {
        "name": bond_name,
        "maturity": maturity_date,
        "purchases": purchases,
        "total_invested": total_invested,
        "total_mkt_gross": total_mkt_gross,
        "total_mkt_net": total_mkt_net,
    }


def calc_real_value(invested, ipca_rate, spread, du):
    """Calculate the real accrued value using the contracted rate.

    Rate = (1 + IPCA) * (1 + spread) - 1, compounded over dias úteis (DU/252).
    """
    yearly_rate = (1 + ipca_rate) * (1 + spread) - 1
    return invested * (1 + yearly_rate) ** (du / 252.0)


def calc_ir_on_real(invested, real_value, ir_rate):
    """Calculate IR tax on the real gain (same bracket as the report)."""
    gain = real_value - invested
    if gain <= 0:
        return 0.0
    return gain * ir_rate


# ============================================================
# === Renda+ projections & stats =============================
# ============================================================

def extract_renda_mais_year(bond_name):
    """Return the year in a Renda+ bond name (e.g. 'Aposentadoria Extra 2035' -> 2035).

    Returns None if the name doesn't match a Renda+ pattern.
    """
    if "renda+" not in bond_name.lower():
        return None
    m = re.search(r"\b(20\d{2})\b", bond_name)
    return int(m.group(1)) if m else None


def renda_mais_conversion_date(year):
    """Conversion date for a Renda+ bond ('data de conversão' = 15 January of the named year)."""
    return date(year, 1, 15)


def monthly_aporte_stats(purchases):
    """Group purchases by calendar month and compute avg/median monthly aporte (BRL)."""
    if not purchases:
        return {"months": 0, "avg": 0.0, "median": 0.0, "total": 0.0,
                "first_month": None, "last_month": None}

    by_month = {}
    for p in purchases:
        key = (p["date"].year, p["date"].month)
        by_month[key] = by_month.get(key, 0.0) + p["invested"]

    keys = sorted(by_month.keys())
    values = sorted(by_month.values())
    n = len(values)
    total = sum(values)
    avg = total / n
    if n % 2 == 1:
        median = values[n // 2]
    else:
        median = (values[n // 2 - 1] + values[n // 2]) / 2

    return {
        "months": n,
        "avg": avg,
        "median": median,
        "total": total,
        "first_month": keys[0],
        "last_month": keys[-1],
    }


BCB_SGS_IPCA_MONTHLY_URL = (
    "https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados?formato=json"
)


def fetch_monthly_ipca_history(years_back, today=None):
    """Fetch monthly IPCA % from BCB SGS series 433 for the last `years_back` calendar years.

    Returns a list of (date, monthly_pct) tuples sorted ascending. monthly_pct is the
    raw percentage (e.g., 0.42 means 0.42% — NOT 0.0042).
    """
    today = today or date.today()
    start_year = today.year - years_back
    url = (
        f"{BCB_SGS_IPCA_MONTHLY_URL}"
        f"&dataInicial=01/01/{start_year}&dataFinal=31/12/{today.year - 1}"
    )
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    out = []
    for entry in resp.json():
        d = datetime.strptime(entry["data"], "%d/%m/%Y").date()
        out.append((d, float(entry["valor"])))
    out.sort(key=lambda t: t[0])
    return out


def compound_monthly_to_yearly(monthly_pcts):
    """Compound a list of monthly IPCA percentages (e.g. 0.42 = 0.42%) into a yearly rate.

    Returns a decimal (e.g. 0.05 means 5%).
    """
    factor = 1.0
    for m in monthly_pcts:
        factor *= 1.0 + m / 100.0
    return factor - 1.0


def yearly_ipca_stats(monthly_series):
    """Aggregate (date, monthly_pct) tuples by calendar year, keep only complete years.

    Returns {'years': [(year, yearly_decimal), ...], 'avg': float, 'median': float}.
    """
    by_year = {}
    for d, m in monthly_series:
        by_year.setdefault(d.year, []).append(m)
    complete = [(y, compound_monthly_to_yearly(ms))
                for y, ms in sorted(by_year.items()) if len(ms) == 12]
    if not complete:
        return {"avg": 0.0, "median": 0.0, "years": []}
    rates = sorted(r for _, r in complete)
    n = len(rates)
    avg = sum(rates) / n
    median = rates[n // 2] if n % 2 == 1 else (rates[n // 2 - 1] + rates[n // 2]) / 2
    return {"avg": avg, "median": median, "years": complete}


def average_spread(purchases):
    """Invested-weighted average contracted spread across purchases."""
    total_inv = sum(p["invested"] for p in purchases)
    if total_inv == 0:
        return 0.0
    return sum(p["invested"] * p["spread"] for p in purchases) / total_inv


def future_aporte_future_value(monthly, yearly_rate, n_months):
    """FV of an ordinary monthly annuity at the given yearly compound rate.

    monthly: BRL deposited at the end of each month.
    yearly_rate: decimal yearly rate (e.g. 0.065).
    n_months: number of monthly deposits.
    """
    if n_months <= 0 or monthly == 0.0:
        return 0.0
    if yearly_rate == 0.0:
        return monthly * n_months
    i = (1 + yearly_rate) ** (1 / 12) - 1
    return monthly * ((1 + i) ** n_months - 1) / i


def project_real_value_at_conversion(
    purchases, conversion_date, today, future_monthly_aporte, avg_future_spread
):
    """Project the real (today's reais) accumulated value at the conversion date.

    Each existing purchase grows at its own contracted spread from its purchase date
    to the conversion date (calendar-year approximation, not DU/252 — the precision
    of a 10–30y projection doesn't warrant the business-day count). Future aportes
    accumulate via the standard ordinary-annuity FV formula at avg_future_spread,
    starting from `today` until the conversion date.
    """
    existing = 0.0
    for p in purchases:
        years = (conversion_date - p["date"]).days / 365.25
        if years < 0:
            years = 0
        existing += p["invested"] * (1 + p["spread"]) ** years

    n_months = max(
        0,
        (conversion_date.year - today.year) * 12 + (conversion_date.month - today.month),
    )
    future = future_aporte_future_value(future_monthly_aporte, avg_future_spread, n_months)
    return existing + future


RENDA_MAIS_N_PAYOUT_MONTHS = 240
RENDA_MAIS_PAYOUT_IR_RATE = 0.15  # always 15% in payout phase (>720d)


def inflate_to_nominal(real_value, yearly_ipca, start, end):
    """Inflate a real BRL value (in `start`-date reais) to nominal at `end`."""
    years = (end - start).days / 365.25
    if years <= 0:
        return real_value
    return real_value * (1 + yearly_ipca) ** years


def renda_mais_payout(
    real_value_at_conversion,
    total_invested,
    conversion_date,
    maturity_date,
    yearly_ipca_for_inflation,
    today,
):
    """Compute the monthly payout during the 240-month Renda+ payout window.

    Returns a dict with real (today's reais) and nominal (first/last payment) values,
    gross and net of 15% IR on the gain portion. Cost basis is split evenly across
    240 payments, so per-payment gain fraction = (V_c - cost) / V_c (in real terms).
    """
    n = RENDA_MAIS_N_PAYOUT_MONTHS
    real_monthly_gross = real_value_at_conversion / n

    gain = max(0.0, real_value_at_conversion - total_invested)
    gain_fraction = gain / real_value_at_conversion if real_value_at_conversion > 0 else 0
    real_monthly_ir = real_monthly_gross * gain_fraction * RENDA_MAIS_PAYOUT_IR_RATE
    real_monthly_net = real_monthly_gross - real_monthly_ir

    nominal_first_gross = inflate_to_nominal(
        real_monthly_gross, yearly_ipca_for_inflation, today, conversion_date
    )
    nominal_first_net = inflate_to_nominal(
        real_monthly_net, yearly_ipca_for_inflation, today, conversion_date
    )
    nominal_last_gross = inflate_to_nominal(
        real_monthly_gross, yearly_ipca_for_inflation, today, maturity_date
    )
    nominal_last_net = inflate_to_nominal(
        real_monthly_net, yearly_ipca_for_inflation, today, maturity_date
    )

    return {
        "n_months": n,
        "real_monthly_gross": real_monthly_gross,
        "real_monthly_net": real_monthly_net,
        "real_monthly_ir": real_monthly_ir,
        "nominal_first_gross": nominal_first_gross,
        "nominal_first_net": nominal_first_net,
        "nominal_last_gross": nominal_last_gross,
        "nominal_last_net": nominal_last_net,
        "total_real_payouts_gross": real_monthly_gross * n,
        "total_real_payouts_net": real_monthly_net * n,
    }


def build_bond_projection(bond, today, aporte_avg, aporte_median, ipca_avg, ipca_median):
    """Compute four projection scenarios per Renda+ bond.

    Returns {'is_renda_mais', 'conversion_date', 'maturity_date', 'avg_spread',
             'scenarios': {scenario_key: {real_value_at_conversion,
                                          nominal_value_at_conversion,
                                          total_invested_through_conversion,
                                          payout: {...}}}}.
    Non-Renda+ bonds return scenarios={}.
    """
    year = extract_renda_mais_year(bond["name"])
    if year is None:
        return {"is_renda_mais": False, "scenarios": {}}

    conversion = renda_mais_conversion_date(year)
    maturity = bond["maturity"]
    avg_spr = average_spread(bond["purchases"])

    n_future_months = max(
        0, (conversion.year - today.year) * 12 + (conversion.month - today.month)
    )

    scenarios = {}
    for ap_label, ap_value in [("avg_aporte", aporte_avg), ("median_aporte", aporte_median)]:
        for ip_label, ip_value in [("avg_ipca", ipca_avg), ("median_ipca", ipca_median)]:
            key = f"{ap_label}_{ip_label}"
            real_vc = project_real_value_at_conversion(
                bond["purchases"], conversion, today,
                future_monthly_aporte=ap_value, avg_future_spread=avg_spr,
            )
            nominal_vc = inflate_to_nominal(real_vc, ip_value, today, conversion)
            total_invested = bond["total_invested"] + ap_value * n_future_months
            payout = renda_mais_payout(
                real_value_at_conversion=real_vc,
                total_invested=total_invested,
                conversion_date=conversion,
                maturity_date=maturity,
                yearly_ipca_for_inflation=ip_value,
                today=today,
            )
            scenarios[key] = {
                "real_value_at_conversion": real_vc,
                "nominal_value_at_conversion": nominal_vc,
                "total_invested_through_conversion": total_invested,
                "future_aporte_monthly": ap_value,
                "ipca_yearly": ip_value,
                "payout": payout,
            }

    return {
        "is_renda_mais": True,
        "conversion_date": conversion,
        "maturity_date": maturity,
        "avg_spread": avg_spr,
        "scenarios": scenarios,
    }


def fmt(value):
    """Format a number as Brazilian currency string."""
    return f"R$ {value:>14,.2f}"


def pct(value):
    """Format a percentage."""
    return f"{value:>7.2f}%"


def write_xlsx(bonds, ipca_rate, today, holidays, aporte_stats, ipca_history_stats, projections):
    """Generate a styled output.xlsx with Detail and Summary sheets."""
    wb = openpyxl.Workbook()

    # -- Styles (Tesouro Direto palette) --
    CURR = '"R$ "#,##0.00'
    DATE_FMT = "dd/mm/yyyy"

    # RendA+ purple gradient
    purple_dark = PatternFill("solid", fgColor="7B1FA2")
    purple_light = PatternFill("solid", fgColor="F3E5F5")
    # Navy brand
    navy = PatternFill("solid", fgColor="0A1F44")
    # Neutrals
    soft_gray = PatternFill("solid", fgColor="F5F7FA")
    slate_border = PatternFill("solid", fgColor="E2E8F0")
    # Accents
    gold = PatternFill("solid", fgColor="FFF8E1")
    green_light = PatternFill("solid", fgColor="E8F5E9")

    white_bold = Font(bold=True, color="FFFFFF", size=11, name="Aptos")
    hdr_font = Font(bold=True, size=10, color="0A1F44", name="Aptos")
    bold = Font(bold=True, size=10, color="1E293B", name="Aptos")
    normal = Font(size=10, color="1E293B", name="Aptos")
    muted = Font(size=10, color="64748B", name="Aptos")
    title_font = Font(bold=True, size=14, color="7B1FA2", name="Aptos")
    subtitle_font = Font(size=11, color="64748B", name="Aptos")

    thin = Border(
        left=Side("thin", color="E2E8F0"),
        right=Side("thin", color="E2E8F0"),
        top=Side("thin", color="E2E8F0"),
        bottom=Side("thin", color="E2E8F0"),
    )

    left = Alignment(horizontal="left")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    def fill_range(ws, r, c_start, c_end, fill):
        for c in range(c_start, c_end + 1):
            ws.cell(r, c).fill = fill

    # ================================================================
    # Detail sheet
    # ================================================================
    ws = wb.active
    ws.title = "Detalhe"

    for i, w in enumerate([14, 8, 18, 16, 8, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title block
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row, 1, "Tesouro Direto - Calculadora de Valor Real")
    c.font = title_font
    c.alignment = left
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row, 1, f"Data: {today.strftime('%d/%m/%Y')}")
    c.font = subtitle_font
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row, 1, f"Taxa IPCA atual (anual): {ipca_rate * 100:.2f}%")
    c.font = subtitle_font
    row += 2

    for bond in bonds:
        maturity_str = bond["maturity"].strftime("%d/%m/%Y")

        # Bond header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row, 1, f"{bond['name']}  |  Vencimento: {maturity_str}")
        c.font = white_bold
        c.fill = purple_dark
        c.alignment = left
        fill_range(ws, row, 2, 8, purple_dark)
        row += 1

        # Column headers
        for ci, h in enumerate(
            [
                "Data",
                "Qtd",
                "Investido",
                "Taxa",
                "DU",
                "Valor Real",
                "Real Líq.",
                "Mkt Líq.",
            ],
            1,
        ):
            c = ws.cell(row, ci, h)
            c.font = hdr_font
            c.fill = purple_light
            c.border = thin
            c.alignment = right if ci > 1 else left
        row += 1

        # Data rows
        bond_invested = 0.0
        bond_real = 0.0
        bond_real_net = 0.0

        for idx, p in enumerate(bond["purchases"]):
            du = count_business_days(p["date"], today, holidays)
            real = calc_real_value(p["invested"], ipca_rate, p["spread"], du)
            ir = calc_ir_on_real(p["invested"], real, p["ir_rate"])
            rnet = real - ir - p["b3_fee"]

            bond_invested += p["invested"]
            bond_real += real
            bond_real_net += rnet

            stripe = soft_gray if idx % 2 == 1 else None
            rate_label = f"IPCA+{p['spread'] * 100:.2f}%"

            vals = [
                p["date"],
                p["qty"],
                p["invested"],
                rate_label,
                du,
                real,
                rnet,
                p["mkt_net"],
            ]
            fmts = [DATE_FMT, "0.00", CURR, None, "0", CURR, CURR, CURR]

            for ci, (v, nf) in enumerate(zip(vals, fmts), 1):
                c = ws.cell(row, ci, v)
                c.font = normal
                c.border = thin
                c.alignment = left if ci == 1 else right
                if nf:
                    c.number_format = nf
                if stripe:
                    c.fill = stripe
            row += 1

        # Total row
        tot_vals = [
            "Total",
            None,
            bond_invested,
            None,
            None,
            bond_real,
            bond_real_net,
            bond["total_mkt_net"],
        ]
        tot_fmts = [None, None, CURR, None, None, CURR, CURR, CURR]
        for ci, (v, nf) in enumerate(zip(tot_vals, tot_fmts), 1):
            c = ws.cell(row, ci, v)
            c.font = bold
            c.fill = purple_light
            c.border = thin
            c.alignment = right if ci > 1 else left
            if nf and v is not None:
                c.number_format = nf
        row += 1

        # Gain summary
        real_gain = bond_real - bond_invested
        real_net_gain = bond_real_net - bond_invested
        mkt_net_gain = bond["total_mkt_net"] - bond["total_invested"]
        bond_gap = bond_real_net - bond["total_mkt_net"]

        gains = [
            (
                "Ganho real bruto",
                real_gain,
                real_gain / bond_invested if bond_invested else 0,
                None,
            ),
            (
                "Ganho real líquido",
                real_net_gain,
                real_net_gain / bond_invested if bond_invested else 0,
                None,
            ),
            (
                "Ganho mercado líquido",
                mkt_net_gain,
                mkt_net_gain / bond["total_invested"] if bond["total_invested"] else 0,
                None,
            ),
            ("Deixando na mesa (real líq. - mkt líq.)", bond_gap, None, gold),
        ]

        for label, value, pct_val, special in gains:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            c = ws.cell(row, 1, label)
            c.font = bold
            c.alignment = right
            if special:
                fill_range(ws, row, 1, 8, special)

            c = ws.cell(row, 6, value)
            c.font = bold
            c.number_format = CURR
            c.alignment = right

            if pct_val is not None:
                c = ws.cell(row, 7, pct_val)
                c.font = bold
                c.number_format = "0.00%"
                c.alignment = right
            row += 1

        row += 1  # blank separator

    # ================================================================
    # Summary sheet
    # ================================================================
    ws2 = wb.create_sheet("Resumo")

    for i, w in enumerate([44, 18, 18, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws2.cell(row, 1, "RESUMO DA CARTEIRA")
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Aptos")
    c.fill = navy
    c.alignment = center
    fill_range(ws2, row, 2, 5, navy)
    row += 2

    # Column headers
    for ci, h in enumerate(
        ["Título", "Investido", "Real Líq.", "Mkt Líq.", "Diferença"], 1
    ):
        c = ws2.cell(row, ci, h)
        c.font = hdr_font
        c.fill = purple_light
        c.border = thin
        c.alignment = right if ci > 1 else left
    row += 1

    # Bond rows
    g_invested = 0.0
    g_real = 0.0
    g_real_net = 0.0
    g_mkt_net = 0.0

    for idx, bond in enumerate(bonds):
        b_inv = 0.0
        b_real = 0.0
        b_rnet = 0.0
        for p in bond["purchases"]:
            du = count_business_days(p["date"], today, holidays)
            real = calc_real_value(p["invested"], ipca_rate, p["spread"], du)
            ir = calc_ir_on_real(p["invested"], real, p["ir_rate"])
            b_inv += p["invested"]
            b_real += real
            b_rnet += real - ir - p["b3_fee"]

        b_gap = b_rnet - bond["total_mkt_net"]
        g_invested += b_inv
        g_real += b_real
        g_real_net += b_rnet
        g_mkt_net += bond["total_mkt_net"]

        stripe = soft_gray if idx % 2 == 1 else None
        for ci, v in enumerate(
            [bond["name"], b_inv, b_rnet, bond["total_mkt_net"], b_gap], 1
        ):
            c = ws2.cell(row, ci, v)
            c.font = normal
            c.border = thin
            c.alignment = right if ci > 1 else left
            if ci > 1:
                c.number_format = CURR
            if stripe:
                c.fill = stripe
        row += 1

    # Total row
    g_gap = g_real_net - g_mkt_net
    for ci, v in enumerate(["TOTAL", g_invested, g_real_net, g_mkt_net, g_gap], 1):
        c = ws2.cell(row, ci, v)
        c.font = bold
        c.fill = purple_light
        c.border = thin
        c.alignment = right if ci > 1 else left
        if ci > 1:
            c.number_format = CURR
    row += 2

    # Summary metrics
    g_real_gain = g_real - g_invested
    g_rnet_gain = g_real_net - g_invested
    g_mkt_gain = g_mkt_net - g_invested

    metrics = [
        ("Total investido", g_invested, None),
        ("Valor real (bruto)", g_real, g_real_gain / g_invested if g_invested else 0),
        (
            "Valor real (líq. de IR)",
            g_real_net,
            g_rnet_gain / g_invested if g_invested else 0,
        ),
        (
            "Valor de mercado (líq.)",
            g_mkt_net,
            g_mkt_gain / g_invested if g_invested else 0,
        ),
        ("Diferença (real líq. - mkt)", g_gap, g_gap / g_mkt_net if g_mkt_net else 0),
    ]

    green_bold = Font(bold=True, size=10, color="2E7D32", name="Aptos")

    for label, value, pct_val in metrics:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws2.cell(row, 1, label)
        c.font = bold
        c.alignment = right

        is_real = "real" in label.lower()
        is_gap = "Diferença" in label
        value_font = green_bold if (is_real or is_gap) and value > 0 else bold

        c = ws2.cell(row, 3, value)
        c.font = value_font
        c.number_format = CURR
        c.alignment = right

        if pct_val is not None:
            c = ws2.cell(row, 4, pct_val)
            c.font = value_font
            c.number_format = "0.00%"
            c.alignment = right

        if is_gap:
            fill_range(ws2, row, 1, 5, gold)
        row += 1

    # ================================================================
    # Estatísticas / Projeções (Renda+)
    # ================================================================
    row += 2

    # --- Aporte stats ---
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws2.cell(row, 1, "ESTATÍSTICAS DE APORTES (por mês-calendário)")
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Aptos")
    c.fill = navy
    c.alignment = center
    fill_range(ws2, row, 2, 5, navy)
    row += 1

    aporte_lines = [
        ("Meses com aporte", aporte_stats["months"], None),
        ("Total aportado", aporte_stats["total"], CURR),
        ("Média mensal", aporte_stats["avg"], CURR),
        ("Mediana mensal", aporte_stats["median"], CURR),
    ]
    for label, value, nf in aporte_lines:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws2.cell(row, 1, label)
        c.font = bold
        c.alignment = right
        c = ws2.cell(row, 3, value)
        c.font = bold
        c.alignment = right
        if nf:
            c.number_format = nf
        row += 1
    row += 1

    # --- IPCA history stats ---
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws2.cell(row, 1, "PROJEÇÃO IPCA (últimos 10 anos, BCB SGS 433)")
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Aptos")
    c.fill = navy
    c.alignment = center
    fill_range(ws2, row, 2, 5, navy)
    row += 1

    for label, value in [("IPCA anual médio", ipca_history_stats["avg"]),
                         ("IPCA anual mediano", ipca_history_stats["median"])]:
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws2.cell(row, 1, label)
        c.font = bold
        c.alignment = right
        c = ws2.cell(row, 3, value)
        c.font = bold
        c.number_format = "0.00%"
        c.alignment = right
        row += 1
    row += 1

    # --- Per-bond Renda+ projections ---
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws2.cell(row, 1, "PROJEÇÃO RENDA+ (por título)")
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Aptos")
    c.fill = navy
    c.alignment = center
    fill_range(ws2, row, 2, 5, navy)
    row += 2

    scenario_labels = {
        "avg_aporte_avg_ipca": "Aporte médio + IPCA médio",
        "avg_aporte_median_ipca": "Aporte médio + IPCA mediano",
        "median_aporte_avg_ipca": "Aporte mediano + IPCA médio",
        "median_aporte_median_ipca": "Aporte mediano + IPCA mediano",
    }

    for bond in bonds:
        proj = projections.get(bond["name"])
        if proj is None or not proj["is_renda_mais"]:
            continue

        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        title = (
            f"{bond['name']}  |  Conversão: {proj['conversion_date'].strftime('%d/%m/%Y')}  "
            f"|  Vencimento: {proj['maturity_date'].strftime('%d/%m/%Y')}"
        )
        c = ws2.cell(row, 1, title)
        c.font = white_bold
        c.fill = purple_dark
        c.alignment = left
        fill_range(ws2, row, 2, 5, purple_dark)
        row += 1

        for ci, h in enumerate(
            ["Cenário", "V_c (real, hoje)", "V_c (nominal)", "Mensal real", "Mensal nominal (1ª)"], 1
        ):
            c = ws2.cell(row, ci, h)
            c.font = hdr_font
            c.fill = purple_light
            c.border = thin
            c.alignment = right if ci > 1 else left
        row += 1

        for idx, key in enumerate(
            ["avg_aporte_avg_ipca", "avg_aporte_median_ipca",
             "median_aporte_avg_ipca", "median_aporte_median_ipca"]
        ):
            scen = proj["scenarios"][key]
            stripe = soft_gray if idx % 2 == 1 else None
            vals = [
                scenario_labels[key],
                scen["real_value_at_conversion"],
                scen["nominal_value_at_conversion"],
                scen["payout"]["real_monthly_net"],
                scen["payout"]["nominal_first_net"],
            ]
            fmts = [None, CURR, CURR, CURR, CURR]
            for ci, (v, nf) in enumerate(zip(vals, fmts), 1):
                c = ws2.cell(row, ci, v)
                c.font = normal
                c.border = thin
                c.alignment = right if ci > 1 else left
                if nf:
                    c.number_format = nf
                if stripe:
                    c.fill = stripe
            row += 1

        # Footer note: last-payment nominal for the avg/avg scenario
        avg_avg = proj["scenarios"]["avg_aporte_avg_ipca"]["payout"]
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws2.cell(
            row, 1,
            f"Mensal nominal líq. na 240ª parcela (avg+avg): "
            f"R$ {avg_avg['nominal_last_net']:,.2f}  •  "
            f"IR descontado: 15% sobre o ganho.",
        )
        c.font = muted
        c.alignment = left
        row += 2

    # Save
    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "output.xlsx"
    )
    wb.save(output_path)
    wb.close()


def main():
    today = date.today()

    print(f"Tesouro Direto - Real Value Calculator")
    print(f"Date: {today.strftime('%d/%m/%Y')}")
    print()

    # Fetch IPCA
    try:
        ipca_rate = fetch_ipca_rate()
    except Exception as e:
        print(f"Error fetching IPCA rate: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Current IPCA rate (yearly): {ipca_rate * 100:.2f}%")
    print()

    # Find all xlsx files — needed early to determine year range for holidays
    xlsx_files = sorted(
        [
            os.path.join(REPORTS_DIR, f)
            for f in os.listdir(REPORTS_DIR)
            if f.endswith(".xlsx")
        ]
    )
    if not xlsx_files:
        print("No .xlsx files found in reports/", file=sys.stderr)
        sys.exit(1)

    # Parse all files
    bonds = []
    for fpath in xlsx_files:
        bonds.append(parse_xlsx(fpath))

    # Sort by maturity
    bonds.sort(key=lambda b: b["maturity"])

    # Fetch holidays for the full date range
    earliest = min(p["date"] for b in bonds for p in b["purchases"])
    try:
        holidays = fetch_holidays(earliest.year, today.year)
    except Exception as e:
        print(f"Error fetching holidays: {e}", file=sys.stderr)
        sys.exit(1)

    # --- Aporte stats (per calendar month, across all bonds) ---
    all_purchases = [p for b in bonds for p in b["purchases"]]
    aporte_stats = monthly_aporte_stats(all_purchases)

    # --- Historical IPCA (last 10 calendar years) ---
    try:
        ipca_history = fetch_monthly_ipca_history(years_back=10, today=today)
    except Exception as e:
        print(f"Warning: failed to fetch historical IPCA: {e}", file=sys.stderr)
        ipca_history = []
    ipca_history_stats = yearly_ipca_stats(ipca_history)

    # --- Per-bond Renda+ projections ---
    projections = {}
    for b in bonds:
        projections[b["name"]] = build_bond_projection(
            bond=b,
            today=today,
            aporte_avg=aporte_stats["avg"],
            aporte_median=aporte_stats["median"],
            ipca_avg=ipca_history_stats["avg"],
            ipca_median=ipca_history_stats["median"],
        )

    # Grand totals
    grand_invested = 0.0
    grand_real = 0.0
    grand_real_net = 0.0
    grand_mkt_gross = 0.0
    grand_mkt_net = 0.0

    for bond in bonds:
        maturity_str = bond["maturity"].strftime("%d/%m/%Y")
        print("=" * 100)
        print(f"  {bond['name']}  |  Maturity: {maturity_str}")
        print("=" * 100)
        print(
            f"  {'Date':<12} {'Qty':>6} {'Invested':>14} {'Rate':>14} "
            f"{'DU':>5} {'Real Value':>14} {'Real Net':>14} {'Mkt Net':>14}"
        )
        print("-" * 100)

        bond_invested = 0.0
        bond_real = 0.0
        bond_real_net = 0.0

        for p in bond["purchases"]:
            du = count_business_days(p["date"], today, holidays)
            real = calc_real_value(p["invested"], ipca_rate, p["spread"], du)
            ir_on_real = calc_ir_on_real(p["invested"], real, p["ir_rate"])
            real_net = real - ir_on_real - p["b3_fee"]

            bond_invested += p["invested"]
            bond_real += real
            bond_real_net += real_net

            rate_label = f"IPCA+{p['spread'] * 100:.2f}%"
            print(
                f"  {p['date'].strftime('%d/%m/%Y'):<12} {p['qty']:>6.2f} "
                f"{fmt(p['invested'])} {rate_label:>14} {du:>5} "
                f"{fmt(real)} {fmt(real_net)} {fmt(p['mkt_net'])}"
            )

        real_gain = bond_real - bond_invested
        real_net_gain = bond_real_net - bond_invested
        mkt_net_gain = bond["total_mkt_net"] - bond["total_invested"]

        print("-" * 100)
        print(
            f"  {'Total':<12} {'':>6} {fmt(bond_invested)} {'':>14} {'':>5} "
            f"{fmt(bond_real)} {fmt(bond_real_net)} {fmt(bond['total_mkt_net'])}"
        )
        print()
        print(
            f"  Real gross gain: {fmt(real_gain)}  ({pct(real_gain / bond_invested * 100)})"
        )
        print(
            f"  Real net gain:   {fmt(real_net_gain)}  ({pct(real_net_gain / bond_invested * 100)})"
        )
        print(
            f"  Market net gain: {fmt(mkt_net_gain)}  ({pct(mkt_net_gain / bond['total_invested'] * 100)})"
        )
        print(
            f"  You're leaving on the table (real net - mkt net): {fmt(bond_real_net - bond['total_mkt_net'])}"
        )
        print()

        grand_invested += bond_invested
        grand_real += bond_real
        grand_real_net += bond_real_net
        grand_mkt_gross += bond["total_mkt_gross"]
        grand_mkt_net += bond["total_mkt_net"]

    # Grand summary
    grand_real_gain = grand_real - grand_invested
    grand_real_net_gain = grand_real_net - grand_invested
    grand_mkt_net_gain = grand_mkt_net - grand_invested
    gap = grand_real_net - grand_mkt_net

    print("=" * 100)
    print("  PORTFOLIO SUMMARY")
    print("=" * 100)
    print()
    print(
        f"  {'Bond':<42} {'Invested':>14} {'Real Net':>14} {'Mkt Net':>14} {'Gap':>14}"
    )
    print(f"  {'-' * 98}")

    for bond in bonds:
        bond_real_net = 0.0
        bond_invested = 0.0
        for p in bond["purchases"]:
            du = count_business_days(p["date"], today, holidays)
            real = calc_real_value(p["invested"], ipca_rate, p["spread"], du)
            ir = calc_ir_on_real(p["invested"], real, p["ir_rate"])
            bond_real_net += real - ir - p["b3_fee"]
            bond_invested += p["invested"]
        bond_gap = bond_real_net - bond["total_mkt_net"]
        print(
            f"  {bond['name']:<42} {fmt(bond_invested)} "
            f"{fmt(bond_real_net)} {fmt(bond['total_mkt_net'])} {fmt(bond_gap)}"
        )

    print(f"  {'-' * 98}")
    print(
        f"  {'TOTAL':<42} {fmt(grand_invested)} "
        f"{fmt(grand_real_net)} {fmt(grand_mkt_net)} {fmt(gap)}"
    )
    print()
    print(f"  Total invested:          {fmt(grand_invested)}")
    print(
        f"  Real value (gross):      {fmt(grand_real)}  ({pct(grand_real_gain / grand_invested * 100)})"
    )
    print(
        f"  Real value (net of IR):  {fmt(grand_real_net)}  ({pct(grand_real_net_gain / grand_invested * 100)})"
    )
    print(
        f"  Market value (net):      {fmt(grand_mkt_net)}  ({pct(grand_mkt_net_gain / grand_invested * 100)})"
    )
    print(f"  Gap (real net - mkt):    {fmt(gap)}  ({pct(gap / grand_mkt_net * 100)})")
    print()

    write_xlsx(bonds, ipca_rate, today, holidays,
               aporte_stats=aporte_stats,
               ipca_history_stats=ipca_history_stats,
               projections=projections)


if __name__ == "__main__":
    buf = io.StringIO()
    original_stdout = sys.stdout
    sys.stdout = type(
        "Tee",
        (),
        {
            "write": lambda self, s: (buf.write(s), original_stdout.write(s)),
            "flush": lambda self: original_stdout.flush(),
        },
    )()
    main()
    sys.stdout = original_stdout
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output.txt")
    with open(output_path, "w") as f:
        f.write(buf.getvalue())
